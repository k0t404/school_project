import os
import sqlite3
import openpyxl
from data import db_session
from data.lessons import Lessons
from comm2 import KeyboardData


class Timetable:
    def __init__(self):
        self.days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']

    def export_to_sqlite(self, grade, file_name):
        # Подключение к бд
        db_session.global_init("db/logs.db")
        db_sess = db_session.create_session()

        # Читаем файл и лист1 книги excel
        file_to_read = openpyxl.load_workbook(f'{file_name}.xlsx', data_only=True)
        sheet = file_to_read[f'{int(grade)} класс']
        # Цикл по строкам
        classes = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1) if sheet.cell(1, col).value]
        day = None
        # Цикл по строчкам
        for row in range(2, 42):
            count_classes = 0
            if str(sheet.cell(row, 2).value) == '1':
                day = sheet.cell(row, 1).value
            # Обновление списка
            data = []
            count_columns = 0
            # Цикл по столбцам
            for col in range(2, (len(classes) + 1) * 3):
                count_columns += 1
                # value содержит значение ячейки с координатами row col
                value = sheet.cell(row, col).value
                # Список который мы потом будем добавлять
                data.append(value)
                if count_columns == 3:
                    data.append(classes[count_classes])
                    count_classes += 1
                    count_columns = 0
                    data.append(day)
                    # Вставка данных в поля таблицы
                    lessons = Lessons()
                    lessons.lesson_pos = data[0]
                    lessons.lesson = data[1]
                    lessons.cabinet = data[2]
                    lessons.class_letter = data[3]
                    lessons.day = data[4]
                    db_sess.add(lessons)

                    data = []

        # сохраняем изменения и закрываем соединение
        db_sess.commit()

    def clear_base(self):
        # Очистка базы sqlite

        # Получаем текущую папку проекта
        prj_dir = os.path.abspath(os.path.curdir)

        # Имя базы
        base_name = 'timetable.db'

        connect = sqlite3.connect(prj_dir + '/db/' + base_name)
        cursor = connect.cursor()

        # Запись в базу, сохранение и закрытие соединения
        cursor.execute("DELETE FROM timetable")
        connect.commit()
        connect.close()

    def process_control(self, *grade_start_end, file_name, wreck_apart=False):
        if wreck_apart:
            pass
        elif grade_start_end:
            grade_start, grade_end = grade_start_end[0]
            for grade in range(int(grade_start), int(grade_end) + 1):
                self.export_to_sqlite(grade, file_name)
